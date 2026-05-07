# ============================================================
# JPSI Futures High / Low Forecast Model
# John Stewart & Partners
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
from datetime import datetime
import plotly.graph_objects as go

# ─── PAGE CONFIG ────────────────────────────────────────────
st.set_page_config(
    page_title="JPSI High/Low Futures Model",
    page_icon="🌽",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── COLORS ─────────────────────────────────────────────────
DM_BG       = "#0d1210"
DM_SURFACE  = "#141c18"
DM_SURFACE2 = "#1a2620"
DM_BORDER   = "#253328"
DM_TEXT     = "#e8ede9"
DM_MUTED    = "#7a9485"
JSA_GREEN   = "#5e7164"
JSA_LT      = "#8db89a"
COL_HIGH    = "#8db89a"
COL_LOW     = "#6fa8c4"
COL_GOLD    = "#c4b456"
COL_PURPLE  = "#9b89c4"
COL_RED     = "#e07070"

# ─── CSS ────────────────────────────────────────────────────
st.markdown(f"""
<style>
html,body,[data-testid="stAppViewContainer"]{{background:{DM_BG};color:{DM_TEXT};}}
[data-testid="stSidebar"]{{display:none;}}
#MainMenu,footer,header{{visibility:hidden;}}
.stDeployButton{{display:none;}}

.hl-row-label{{
  font-size:0.72rem;text-transform:uppercase;letter-spacing:0.08em;
  color:{DM_MUTED};margin-bottom:6px;padding-top:4px;
}}
.hl-card{{
  background:{DM_SURFACE};border:1px solid {DM_BORDER};border-radius:12px;
  padding:16px 18px 14px;text-align:center;height:100%;
}}
.hl-card-high{{border-top:4px solid {COL_HIGH};}}
.hl-card-low{{border-top:4px solid {COL_LOW};}}
.hl-contract{{color:{DM_MUTED};font-size:0.65rem;text-transform:uppercase;
  letter-spacing:0.10em;margin-bottom:2px;}}
.hl-label{{color:{DM_MUTED};font-size:0.60rem;text-transform:uppercase;
  letter-spacing:0.08em;margin-bottom:8px;}}
.hl-price-high{{color:{COL_HIGH};font-size:2rem;font-weight:700;line-height:1;}}
.hl-price-low{{color:{COL_LOW};font-size:2rem;font-weight:700;line-height:1;}}
.hl-sub{{color:{DM_MUTED};font-size:0.76rem;margin-top:5px;}}

.sec-hdr{{
  font-size:1.0rem;font-weight:600;color:{DM_TEXT};
  border-left:4px solid {JSA_LT};padding-left:10px;
  margin:1.4rem 0 0.7rem 0;
}}
.sec-div{{
  border-bottom:1px solid {DM_BORDER};margin:6px 0 12px 0;
  color:{DM_MUTED};font-size:0.66rem;text-transform:uppercase;letter-spacing:0.10em;
  padding-bottom:4px;
}}
.ratio-badge{{
  background:{DM_SURFACE2};border:1px solid {DM_BORDER};border-radius:6px;
  padding:3px 10px;display:inline-block;font-size:0.82rem;
  color:{DM_TEXT};margin-top:4px;
}}
.tbl-header{{
  color:{DM_MUTED};font-size:0.68rem;text-transform:uppercase;
  letter-spacing:0.07em;margin:0.8rem 0 0.3rem;
}}
.indicated-box{{
  background:{DM_SURFACE2};border:1px solid {DM_BORDER};border-radius:8px;
  padding:10px 14px;margin-bottom:10px;font-size:0.88rem;
}}

/* Input tab card styling */
.input-card{{
  background:{DM_SURFACE};border:1px solid {DM_BORDER};border-radius:12px;
  padding:20px 22px 18px;margin-bottom:4px;
}}
.input-card-title{{
  font-size:1.0rem;font-weight:700;color:{DM_TEXT};
  margin-bottom:14px;padding-bottom:10px;
  border-bottom:1px solid {DM_BORDER};
}}

.stTabs [data-baseweb="tab-list"]{{
  background:{DM_SURFACE};border-radius:10px;
  padding:6px 8px;gap:6px;border:1px solid {DM_BORDER};
}}
.stTabs [data-baseweb="tab"]{{
  color:{DM_MUTED};font-size:0.95rem;font-weight:600;
  padding:9px 22px;border-radius:7px;
}}
.stTabs [data-baseweb="tab"]:hover{{background:{DM_SURFACE2};color:{DM_TEXT};}}
.stTabs [aria-selected="true"]{{color:#fff !important;background:{JSA_GREEN} !important;}}
.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"]{{display:none !important;}}

div[data-testid="stDataFrame"]{{
  background:{DM_SURFACE};border-radius:8px;border:1px solid {DM_BORDER};
}}
</style>
""", unsafe_allow_html=True)


# ─── DATA LOADING ───────────────────────────────────────────
DATA_PATH = Path(__file__).parent / "data" / "High Low Model.xlsx"

# External price history files — update the primary paths below as needed.
# Fallback: app also checks the local data/ folder (drop copies there for
# Streamlit Cloud or when the network path is unavailable).
_CORN_PRIMARY = Path(r"C:\Users\KoltenPostin\John Stewart and Associates\JSA - Documents\Research Analyst\Reports\Projects\Python Codes\Prices & Spreads\Corn_Futures_History.xlsx")
_SOY_PRIMARY  = Path(r"C:\Users\KoltenPostin\John Stewart and Associates\JSA - Documents\Research Analyst\Reports\Projects\Python Codes\Prices & Spreads\Soybean_Futures_History.xlsx")
_CORN_LOCAL   = Path(__file__).parent / "data" / "Corn_Futures_History.xlsx"
_SOY_LOCAL    = Path(__file__).parent / "data" / "Soybean_Futures_History.xlsx"

CORN_PRICES_PATH = _CORN_PRIMARY if _CORN_PRIMARY.exists() else _CORN_LOCAL
SOY_PRICES_PATH  = _SOY_PRIMARY  if _SOY_PRIMARY.exists()  else _SOY_LOCAL


def refresh_price_files():
    """
    Copy the latest price files from the SharePoint-synced path into data/.
    Returns (success: bool, message: str).
    """
    import shutil
    messages = []
    success = True
    for src, dst, label in [
        (_CORN_PRIMARY, _CORN_LOCAL, "Corn"),
        (_SOY_PRIMARY,  _SOY_LOCAL,  "Soybeans"),
    ]:
        if src.exists():
            try:
                shutil.copy2(src, dst)
                messages.append(f"✅ {label}: copied from SharePoint sync folder")
            except Exception as e:
                messages.append(f"❌ {label}: copy failed — {e}")
                success = False
        elif dst.exists():
            messages.append(f"⚠️ {label}: SharePoint path not reachable, keeping existing data/ file")
        else:
            messages.append(f"❌ {label}: no file found at either path")
            success = False
    load_price_history.clear()   # bust the cache so next call re-parses
    return success, "\n\n".join(messages)


def _parse_year(v):
    if v is None:
        return None
    try:
        return int(str(v).strip())
    except Exception:
        return None


def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


@st.cache_data
def load_data():
    wb = openpyxl.load_workbook(DATA_PATH)
    return {
        "cz_low":  _load_cz(wb, "CZ  Low"),
        "cz_high": _load_cz(wb, "CZ High"),
        "cn_low":  _load_cn(wb, "CN Low"),
        "cn_high": _load_cn(wb, "CN High"),
        "sx_low":  _load_sx(wb, "SX Low"),
        "sx_high": _load_sx(wb, "SX High"),
        "sn_low":  _load_sn(wb, "SN Low",  is_low=True),
        "sn_high": _load_sn(wb, "SN High", is_low=False),
    }


def _load_cz(wb, name):
    rows = list(wb[name].iter_rows(values_only=True))
    out = []
    for row in rows:
        yr = _parse_year(row[1])
        if not (yr and 1985 <= yr <= 2030):
            continue
        usage, prod, jan1, price = row[2], row[4], row[8], row[12]
        if not all(_is_num(v) for v in [usage, prod, jan1, price]):
            continue
        out.append(dict(
            year=yr,
            usage=usage,
            production=prod,
            ratio=prod / usage,
            jan1=jan1,
            date=row[10] if isinstance(row[10], datetime) else None,
            price=price,
            price_pct=price / jan1,
            is_est=(row[13] == "Est"),
        ))
    df = pd.DataFrame(out)
    if not df.empty:
        df["section"] = df["ratio"].apply(
            lambda r: "Prod ≥ Use" if r >= 1 else "Prod < Use"
        )
    return df


def _load_cn(wb, name):
    rows = list(wb[name].iter_rows(values_only=True))
    out = []
    for i, row in enumerate(rows):
        if i >= 48:
            break
        yr = _parse_year(row[1])
        if not (yr and 1985 <= yr <= 2030):
            continue
        co_pct, price, jan1 = row[4], row[5], row[9]
        if not all(_is_num(v) for v in [co_pct, price, jan1]):
            continue
        out.append(dict(
            year=yr,
            carryout_pct=co_pct,
            jan1=jan1,
            date=row[7] if isinstance(row[7], datetime) else None,
            price=price,
            price_pct=price / jan1,
            is_est=(row[6] == "Est"),
        ))
    return pd.DataFrame(out)


def _load_sx(wb, name):
    rows = list(wb[name].iter_rows(values_only=True))
    out = []
    section = "Prod ≥ Prev Yr Use"
    for i, row in enumerate(rows):
        if i < 8:
            continue
        rstr = " ".join(str(v) for v in row if v is not None)
        if "Less Than" in rstr:
            section = "Prod < Prev Yr Use"
            continue
        if "Equal to or Greater" in rstr:
            section = "Prod ≥ Prev Yr Use"
            continue
        yr = _parse_year(row[0])
        if not (yr and 1985 <= yr <= 2030):
            continue
        co_pct, jan1, price = row[1], row[3], row[7]
        if not all(_is_num(v) for v in [co_pct, jan1, price]):
            continue
        out.append(dict(
            year=yr,
            carryout_pct=co_pct,
            jan1=jan1,
            date=row[5] if isinstance(row[5], datetime) else None,
            price=price,
            price_pct=price / jan1,
            is_est=(row[8] == "Est"),
            section=section,
        ))
    return pd.DataFrame(out)


def _load_sn(wb, name, is_low):
    rows = list(wb[name].iter_rows(values_only=True))
    out = []
    section = "No Crop Scare"
    yr_idx  = 1 if is_low else 0
    co_idx  = 6 if is_low else 5

    for i, row in enumerate(rows):
        if i < 9:
            continue
        rstr = " ".join(str(v) for v in row if v is not None)
        rl = rstr.lower()
        if "without" in rl and "crop scare" in rl:
            section = "No Crop Scare"
            continue
        if "south america" in rl:
            section = "SA Crop Problem"
            continue
        if "with" in rl and "crop scare" in rl and "without" not in rl:
            section = "Crop Scare"
            continue

        yr = _parse_year(row[yr_idx])
        if not (yr and 1985 <= yr <= 2030):
            continue
        co_pct, price, jan1 = row[co_idx], row[7], row[11]
        if not all(_is_num(v) for v in [co_pct, price, jan1]):
            continue
        out.append(dict(
            year=yr,
            carryout_pct=co_pct,
            jan1=jan1,
            date=row[9] if isinstance(row[9], datetime) else None,
            price=price,
            price_pct=price / jan1,
            is_est=(row[8] == "Est"),
            section=section,
        ))
    return pd.DataFrame(out)


# ─── PRICE HISTORY LOADER ───────────────────────────────────

@st.cache_data
def load_price_history():
    """
    Parse both wide-format futures price files.
    Returns dict {"corn": {ticker: df}, "soy": {ticker: df}} or None if files missing.
    Each df has a DatetimeIndex and columns: Open, High, Low, Close, OI, Volume.
    Prices are in ¢/bu (raw file values).
    """
    if not CORN_PRICES_PATH.exists() or not SOY_PRICES_PATH.exists():
        return None

    def parse_file(path):
        contracts = {}
        xl = pd.ExcelFile(path, engine="openpyxl")
        for sheet in xl.sheet_names:
            raw = pd.read_excel(xl, sheet_name=sheet, header=None)
            # Row 0 = ticker symbols, Row 1 = field headers, Row 2+ = data
            dates = pd.to_datetime(raw.iloc[2:, 0], errors="coerce")
            col = 1
            while col + 5 < len(raw.columns):
                ticker = raw.iloc[0, col]
                if pd.isna(ticker) or str(ticker).strip() == "":
                    col += 7
                    continue
                ticker = str(ticker).strip()
                block = raw.iloc[2:, col:col + 6].copy()
                block.columns = ["Open", "High", "Low", "Close", "OI", "Volume"]
                block.index = dates.values
                block = block[pd.notna(block.index)]
                for c in block.columns:
                    block[c] = pd.to_numeric(block[c], errors="coerce")
                block = block.dropna(subset=["Close"])
                if not block.empty:
                    contracts[ticker] = block
                col += 7
        return contracts

    return {
        "corn": parse_file(CORN_PRICES_PATH),
        "soy":  parse_file(SOY_PRICES_PATH),
    }


def _ticker_year(ticker):
    """Extract 4-digit delivery year from a ticker like ZCZ25 → 2025."""
    suffix = ticker[-2:]
    try:
        y2 = int(suffix)
        return 2000 + y2 if y2 < 60 else 1900 + y2
    except ValueError:
        return None


def build_seasonal_df(contracts, prefix, delivery_year_filter=None):
    """
    For contracts whose ticker starts with `prefix` (e.g. 'ZCZ'),
    filter each contract's data to just dates in its delivery year,
    normalize Close to % of the first trading day of that year,
    and return a long DataFrame with columns:
        year, date, doy (day of year 1-366), close_raw, price_pct
    Optionally restrict to delivery years in `delivery_year_filter` (list of ints).
    """
    rows = []
    for ticker, df in contracts.items():
        if not ticker.startswith(prefix):
            continue
        year = _ticker_year(ticker)
        if year is None:
            continue
        if delivery_year_filter and year not in delivery_year_filter:
            continue

        idx = pd.DatetimeIndex(df.index)
        year_df = df[idx.year == year].copy()
        if len(year_df) < 10:
            continue

        jan1_price = year_df.iloc[0]["Close"]
        if pd.isna(jan1_price) or jan1_price <= 0:
            continue

        for ts, row in year_df.iterrows():
            dt = pd.Timestamp(ts)
            rows.append({
                "year":      year,
                "date":      dt,
                "doy":       dt.timetuple().tm_yday,
                "close_raw": row["Close"],
                "price_pct": row["Close"] / jan1_price * 100,
            })
    return pd.DataFrame(rows)


def make_seasonal_overlay(seasonal_df, current_year, title):
    """
    Seasonal overlay chart: each historical year as a faint line,
    historical average ± 1 std dev band in green, current year in gold.
    X-axis = day of year labelled by month. Y-axis = % of Jan 1.
    """
    fig = go.Figure()

    hist = seasonal_df[seasonal_df["year"] != current_year]
    curr = seasonal_df[seasonal_df["year"] == current_year]

    # Faint historical year lines
    years = sorted(hist["year"].unique())
    for yr in years:
        g = hist[hist["year"] == yr].sort_values("doy")
        fig.add_trace(go.Scatter(
            x=g["doy"], y=g["price_pct"],
            mode="lines",
            line=dict(width=0.7, color="rgba(94,113,100,0.25)"),
            name=str(yr),
            showlegend=False,
            hovertemplate=f"<b>{yr}</b><br>%{{x:.0f}} · %{{y:.1f}}% of Jan 1<extra></extra>",
        ))

    # Average ± 1 SD band
    avg = hist.groupby("doy")["price_pct"].mean()
    std = hist.groupby("doy")["price_pct"].std().fillna(0)
    doys = avg.index.tolist()

    fig.add_trace(go.Scatter(
        x=doys + doys[::-1],
        y=list(avg + std) + list((avg - std).iloc[::-1]),
        fill="toself",
        fillcolor="rgba(94,113,100,0.18)",
        line=dict(width=0),
        name="±1 Std Dev",
        hoverinfo="skip",
    ))
    fig.add_trace(go.Scatter(
        x=doys, y=avg.values,
        mode="lines",
        line=dict(width=2, color=JSA_LT),
        name="Historical Avg",
        hovertemplate="Avg: %{y:.1f}% of Jan 1<extra></extra>",
    ))

    # Current year
    if not curr.empty:
        curr = curr.sort_values("doy")
        fig.add_trace(go.Scatter(
            x=curr["doy"], y=curr["price_pct"],
            mode="lines",
            line=dict(width=2.5, color=COL_GOLD),
            name=f"{current_year} (Current)",
            hovertemplate=f"<b>{current_year}</b><br>%{{x:.0f}} · %{{y:.1f}}% of Jan 1<extra></extra>",
        ))

    # Jan 1 baseline
    fig.add_hline(y=100, line_dash="dot", line_color=DM_MUTED, line_width=1,
                  annotation_text=" Jan 1 = 100%", annotation_font=dict(color=DM_MUTED, size=10))

    # Month tick marks (approx. first day of each month in a non-leap year)
    month_doys  = [1, 32, 60, 91, 121, 152, 182, 213, 244, 274, 305, 335]
    month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    fig.update_layout(
        title=dict(text=title, font=dict(color=DM_TEXT, size=12), x=0),
        plot_bgcolor=DM_SURFACE2,
        paper_bgcolor=DM_SURFACE2,
        font=dict(color=DM_TEXT, family="Arial"),
        xaxis=dict(
            title="Month",
            tickvals=month_doys, ticktext=month_names,
            gridcolor=DM_BORDER, color=DM_MUTED, zeroline=False,
            range=[1, 366],
        ),
        yaxis=dict(
            title="% of Jan 1 Price",
            gridcolor=DM_BORDER, color=DM_MUTED,
            zeroline=False, ticksuffix="%",
        ),
        legend=dict(bgcolor=DM_SURFACE, bordercolor=DM_BORDER, borderwidth=1,
                    font=dict(color=DM_TEXT, size=10)),
        height=480,
        margin=dict(l=55, r=20, t=45, b=50),
    )
    return fig


# ─── COMPUTATION HELPERS ────────────────────────────────────

def rank_df(df, ratio_col, current_ratio, current_jan1):
    """Filter est rows, rank by proximity to current ratio, add indicated price."""
    d = df[~df["is_est"]].copy()
    d["distance"] = (d[ratio_col] - current_ratio).abs()
    d = d.sort_values("distance").reset_index(drop=True)
    d.index = d.index + 1
    d["indicated"] = (d["price_pct"] * current_jan1).round(2)
    return d


def headline(df, ratio_col, current_ratio, current_jan1):
    """Return (median_indicated_price, median_pct, top5_indicated_price)."""
    d = rank_df(df, ratio_col, current_ratio, current_jan1)
    if d.empty:
        return None, None, None
    med_pct  = d["price_pct"].median()
    top5_pct = d.head(5)["price_pct"].median()
    return round(med_pct * current_jan1, 2), med_pct, round(top5_pct * current_jan1, 2)


def fmt_pct(p):
    return f"{p*100:.1f}%" if p is not None else "—"


def fmt_price(p):
    """Display price in $/bu (inputs are in ¢/bu, divide by 100)."""
    return f"${p/100:.2f}/bu" if p is not None else "—"



# ─── SECTION EMOJI MAP ──────────────────────────────────────

SECTION_EMOJI = {
    "Prod ≥ Use":          "🟢",
    "Prod < Use":          "🔴",
    "Prod ≥ Prev Yr Use":  "🟢",
    "Prod < Prev Yr Use":  "🔴",
    "No Crop Scare":       "🟢",
    "SA Crop Problem":     "🟡",
    "Crop Scare":          "🔴",
}


# ─── TABLE BUILDER ──────────────────────────────────────────

def build_display_table(df, ratio_col, ratio_label, current_ratio, current_jan1,
                         price_col_label, current_section=None):
    d = rank_df(df, ratio_col, current_ratio, current_jan1)
    if d.empty:
        return pd.DataFrame()

    out = pd.DataFrame()
    out["Rank"]              = d.index
    out["Year"]              = d["year"].astype(int)

    if "section" in d.columns:
        out["Category"] = d["section"].apply(
            lambda s: f"{SECTION_EMOJI.get(s, '⚪')} {s}"
        )
        if current_section:
            out["Cur Yr"] = d["section"].apply(
                lambda s: "★" if s == current_section else ""
            )

    out[ratio_label]          = (d[ratio_col] * 100).round(2).astype(str) + "%"
    out["Dist. from Current"] = (d["distance"] * 100).round(3).astype(str) + "%"
    out["Jan 1 (¢/bu)"]      = d["jan1"].round(2)
    out[price_col_label]      = d["price"].round(2)
    out["% of Jan 1"]         = (d["price_pct"] * 100).round(1).astype(str) + "%"
    out["Indicated ($/bu)"]   = (d["indicated"] / 100).round(4)
    out["Date of H/L"]        = d["date"].apply(
        lambda x: x.strftime("%m/%d/%Y") if (pd.notna(x) and isinstance(x, datetime)) else "—"
    )
    return out


# ─── SCATTER CHART ──────────────────────────────────────────

SECTION_COLORS_LOW = {
    "Prod ≥ Use":          COL_LOW,
    "Prod < Use":          COL_GOLD,
    "Prod ≥ Prev Yr Use":  COL_LOW,
    "Prod < Prev Yr Use":  COL_GOLD,
    "No Crop Scare":       COL_LOW,
    "SA Crop Problem":     COL_GOLD,
    "Crop Scare":          COL_RED,
    "All":                 COL_LOW,
}
SECTION_COLORS_HIGH = {
    "Prod ≥ Use":          COL_HIGH,
    "Prod < Use":          COL_PURPLE,
    "Prod ≥ Prev Yr Use":  COL_HIGH,
    "Prod < Prev Yr Use":  COL_PURPLE,
    "No Crop Scare":       COL_HIGH,
    "SA Crop Problem":     COL_PURPLE,
    "Crop Scare":          COL_RED,
    "All":                 COL_HIGH,
}


def make_scatter(df_low, df_high, ratio_col, ratio_label, current_ratio, title):
    fig = go.Figure()

    for df, lbl, cmap, sym in [
        (df_low,  "Low",  SECTION_COLORS_LOW,  "circle"),
        (df_high, "High", SECTION_COLORS_HIGH, "diamond"),
    ]:
        d = df[~df["is_est"]].copy()
        if "section" in d.columns:
            groups = [(s, d[d["section"] == s]) for s in d["section"].unique()]
        else:
            groups = [("All", d)]

        for sec, sd in groups:
            if sd.empty:
                continue
            color = cmap.get(sec, DM_MUTED)
            fig.add_trace(go.Scatter(
                x=sd[ratio_col] * 100,
                y=sd["price_pct"] * 100,
                mode="markers+text",
                text=sd["year"].astype(str),
                textposition="top center",
                textfont=dict(size=8, color=DM_MUTED),
                marker=dict(size=9, color=color, symbol=sym,
                            line=dict(width=1, color=DM_BORDER)),
                name=f"{lbl} — {sec}",
                hovertemplate=(
                    f"<b>%{{text}}</b> ({lbl})<br>"
                    f"{ratio_label}: %{{x:.2f}}%<br>"
                    "Price % of Jan 1: %{y:.1f}%<br>"
                    "<extra></extra>"
                ),
            ))

    fig.add_vline(
        x=current_ratio * 100,
        line_color=JSA_LT, line_dash="dash", line_width=2,
        annotation_text=f" Current: {current_ratio*100:.2f}%",
        annotation_position="top right",
        annotation_font=dict(color=JSA_LT, size=11),
    )

    fig.update_layout(
        title=dict(text=title, font=dict(color=DM_TEXT, size=12), x=0),
        plot_bgcolor=DM_SURFACE2,
        paper_bgcolor=DM_SURFACE2,
        font=dict(color=DM_TEXT, family="Arial"),
        xaxis=dict(title=ratio_label, gridcolor=DM_BORDER, color=DM_MUTED,
                   ticksuffix="%", zeroline=False),
        yaxis=dict(title="Price as % of Jan 1", gridcolor=DM_BORDER, color=DM_MUTED,
                   ticksuffix="%", zeroline=False),
        legend=dict(bgcolor=DM_SURFACE, bordercolor=DM_BORDER, borderwidth=1,
                    font=dict(color=DM_TEXT, size=10)),
        height=430,
        margin=dict(l=55, r=20, t=40, b=50),
    )
    return fig


# ─── HEADLINE TILE BUILDER ──────────────────────────────────

def headline_tile(contract, label, price, pct, top5, kind):
    cls        = "high" if kind == "high" else "low"
    price_html = fmt_price(price)
    pct_html   = f"Median: {fmt_pct(pct)} of Jan 1" if pct else "—"
    top5_html  = f"Top 5 Comps: {fmt_price(top5)}" if top5 else ""
    return f"""
<div class="hl-card hl-card-{cls}">
  <div class="hl-contract">{contract}</div>
  <div class="hl-label">{label}</div>
  <div class="hl-price-{cls}">{price_html}</div>
  <div class="hl-sub">{pct_html}</div>
  <div class="hl-sub">{top5_html}</div>
</div>"""


# ─── READ INPUTS FROM SESSION STATE (with defaults) ─────────
# Inputs live in the Assumptions tab but are consumed above it in the
# Overview tab. Reading from session_state here means any change in the
# Assumptions tab triggers a full re-run with updated values everywhere.

cz_jan1      = st.session_state.get("cz_jan1",      458.50)
cz_prod      = st.session_state.get("cz_prod",    15979.0)
cz_use       = st.session_state.get("cz_use",     16120.0)
cn_jan1      = st.session_state.get("cn_jan1",      452.00)
cn_co_raw    = st.session_state.get("cn_co",         12.9)
sx_jan1      = st.session_state.get("sx_jan1",    1062.75)
sx_co_raw    = st.session_state.get("sx_co",         28.4)
sx_section   = st.session_state.get("sx_section", "Prod ≥ Prev Yr Use")
sn_jan1      = st.session_state.get("sn_jan1",    1072.00)
sn_co_raw    = st.session_state.get("sn_co",         29.5)
sn_section   = st.session_state.get("sn_section", "No Crop Scare")

cz_ratio  = cz_prod / cz_use if cz_use else 0.0
cn_co_pct = cn_co_raw / 100
sx_co_pct = sx_co_raw / 100
sn_co_pct = sn_co_raw / 100
cz_section = "Prod ≥ Use" if cz_ratio >= 1 else "Prod < Use"


# ─── LOAD DATA ──────────────────────────────────────────────
try:
    D = load_data()
except FileNotFoundError:
    st.error(
        "📂 Data file not found. Make sure `data/High Low Model.xlsx` "
        "is in the same folder as `app.py`."
    )
    st.stop()
except Exception as e:
    st.error(f"Error loading data: {e}")
    st.stop()


# ─── HEADLINE COMPUTATIONS ──────────────────────────────────
cz_hl_low,  cz_lp,  cz_t5_low  = headline(D["cz_low"],  "ratio",        cz_ratio,  cz_jan1)
cz_hl_high, cz_hp,  cz_t5_high = headline(D["cz_high"], "ratio",        cz_ratio,  cz_jan1)
cn_hl_low,  cn_lp,  cn_t5_low  = headline(D["cn_low"],  "carryout_pct", cn_co_pct, cn_jan1)
cn_hl_high, cn_hp,  cn_t5_high = headline(D["cn_high"], "carryout_pct", cn_co_pct, cn_jan1)
sx_hl_low,  sx_lp,  sx_t5_low  = headline(D["sx_low"],  "carryout_pct", sx_co_pct, sx_jan1)
sx_hl_high, sx_hp,  sx_t5_high = headline(D["sx_high"], "carryout_pct", sx_co_pct, sx_jan1)
sn_hl_low,  sn_lp,  sn_t5_low  = headline(D["sn_low"],  "carryout_pct", sn_co_pct, sn_jan1)
sn_hl_high, sn_hp,  sn_t5_high = headline(D["sn_high"], "carryout_pct", sn_co_pct, sn_jan1)


# ─── SEASONALITY CHART ──────────────────────────────────────

MONTH_LABELS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def make_seasonality_chart(df_low, df_high, title):
    """Bar chart: how often does the annual high/low fall in each calendar month."""
    def month_counts(df):
        dates = df[~df["is_est"]]["date"]
        months = dates.apply(lambda x: x.month if (pd.notna(x) and isinstance(x, datetime)) else None).dropna()
        counts = [int((months == m).sum()) for m in range(1, 13)]
        return counts

    low_counts  = month_counts(df_low)
    high_counts = month_counts(df_high)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=MONTH_LABELS, y=low_counts,
        name="Annual Low",
        marker_color=COL_LOW,
        marker_line=dict(width=0),
    ))
    fig.add_trace(go.Bar(
        x=MONTH_LABELS, y=high_counts,
        name="Annual High",
        marker_color=COL_HIGH,
        marker_line=dict(width=0),
    ))
    fig.update_layout(
        title=dict(text=title, font=dict(color=DM_TEXT, size=12), x=0),
        barmode="group",
        plot_bgcolor=DM_SURFACE2,
        paper_bgcolor=DM_SURFACE2,
        font=dict(color=DM_TEXT, family="Arial"),
        xaxis=dict(title="Month", gridcolor=DM_BORDER, color=DM_MUTED, zeroline=False),
        yaxis=dict(title="Number of Years", gridcolor=DM_BORDER, color=DM_MUTED,
                   zeroline=False, dtick=1),
        legend=dict(bgcolor=DM_SURFACE, bordercolor=DM_BORDER, borderwidth=1,
                    font=dict(color=DM_TEXT, size=10)),
        height=320,
        margin=dict(l=50, r=20, t=40, b=50),
    )
    return fig


# ─── PAGE HEADER ────────────────────────────────────────────
LOGO_URL = "https://www.jpsi.com/wp-content/themes/gate39media/img/logo-white.png"

hdr_l, hdr_r = st.columns([1, 4])
with hdr_l:
    st.image(LOGO_URL, width=160)
with hdr_r:
    st.markdown(
        f"""
        <div style="display:flex;align-items:baseline;gap:14px;margin-top:8px;">
          <h1 style="margin:0;font-size:1.55rem;font-weight:700;color:{DM_TEXT};">
            Futures High / Low Forecast Model
          </h1>
          <span style="color:{DM_MUTED};font-size:0.85rem;">
            John Stewart &amp; Partners &nbsp;·&nbsp; Historical Analog Pricing
          </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)

# ─── LOAD PRICE HISTORY ─────────────────────────────────────
PH = load_price_history()   # None if files not on this machine

# ─── TOP-LEVEL TABS ─────────────────────────────────────────
tab_ov, tab_inp, tab_cz, tab_cn, tab_sx, tab_sn, tab_seas = st.tabs([
    "📌  Overview",
    "⚙️  Assumptions",
    "🌽  Dec Corn (CZ)",
    "🌽  Jul Corn (CN)",
    "🫘  Nov Soybeans (SX)",
    "🫘  Jul Soybeans (SN)",
    "📈  Seasonals",
])


# ══════════════════════════════════════════════════════════════
# OVERVIEW TAB — headline price tiles
# ══════════════════════════════════════════════════════════════
with tab_ov:
    st.markdown(
        '<div class="sec-hdr">📌 Indicated Price Range — Current Marketing Year</div>',
        unsafe_allow_html=True,
    )

    # Row 1: Lows
    st.markdown(
        '<div class="hl-row-label">📉 Indicated Lows (Median of All Historical Years × Current Jan 1 Price)</div>',
        unsafe_allow_html=True,
    )
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(headline_tile("CZ — Dec Corn",  "Jan–Dec Indicated Low",  cz_hl_low,  cz_lp,  cz_t5_low,  "low"),  unsafe_allow_html=True)
    c2.markdown(headline_tile("CN — Jul Corn",  "Jan–Jul Indicated Low",  cn_hl_low,  cn_lp,  cn_t5_low,  "low"),  unsafe_allow_html=True)
    c3.markdown(headline_tile("SX — Nov Beans", "Jan–Nov Indicated Low",  sx_hl_low,  sx_lp,  sx_t5_low,  "low"),  unsafe_allow_html=True)
    c4.markdown(headline_tile("SN — Jul Beans", "Jan–Jul Indicated Low",  sn_hl_low,  sn_lp,  sn_t5_low,  "low"),  unsafe_allow_html=True)

    st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)

    # Row 2: Highs
    st.markdown(
        '<div class="hl-row-label">📈 Indicated Highs (Median of All Historical Years × Current Jan 1 Price)</div>',
        unsafe_allow_html=True,
    )
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(headline_tile("CZ — Dec Corn",  "Jan–Dec Indicated High", cz_hl_high, cz_hp, cz_t5_high, "high"), unsafe_allow_html=True)
    c2.markdown(headline_tile("CN — Jul Corn",  "Jan–Jul Indicated High", cn_hl_high, cn_hp, cn_t5_high, "high"), unsafe_allow_html=True)
    c3.markdown(headline_tile("SX — Nov Beans", "Jan–Nov Indicated High", sx_hl_high, sx_hp, sx_t5_high, "high"), unsafe_allow_html=True)
    c4.markdown(headline_tile("SN — Jul Beans", "Jan–Jul Indicated High", sn_hl_high, sn_hp, sn_t5_high, "high"), unsafe_allow_html=True)

    st.markdown('<div style="height:16px;"></div>', unsafe_allow_html=True)

    # Current assumptions summary strip
    st.markdown('<div class="sec-hdr">📋 Current Assumptions Summary</div>', unsafe_allow_html=True)
    s1, s2, s3, s4 = st.columns(4)
    with s1:
        st.markdown(
            f'<div class="hl-card" style="text-align:left;">'
            f'<div class="hl-contract" style="margin-bottom:8px;">🌽 Dec Corn (CZ)</div>'
            f'<div style="font-size:0.82rem;color:{DM_MUTED};">Jan 1 Price</div>'
            f'<div style="font-size:1.1rem;font-weight:600;color:{DM_TEXT};">${cz_jan1/100:.2f}/bu</div>'
            f'<div style="font-size:0.82rem;color:{DM_MUTED};margin-top:6px;">Prod/Use</div>'
            f'<div style="font-size:1.0rem;color:{DM_TEXT};">{cz_ratio*100:.2f}%'
            f' &nbsp;<span style="font-size:0.75rem;color:{JSA_LT if cz_ratio>=1 else COL_RED};">'
            f'({cz_section})</span></div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    with s2:
        st.markdown(
            f'<div class="hl-card" style="text-align:left;">'
            f'<div class="hl-contract" style="margin-bottom:8px;">🌽 Jul Corn (CN)</div>'
            f'<div style="font-size:0.82rem;color:{DM_MUTED};">Jan 1 Price</div>'
            f'<div style="font-size:1.1rem;font-weight:600;color:{DM_TEXT};">${cn_jan1/100:.2f}/bu</div>'
            f'<div style="font-size:0.82rem;color:{DM_MUTED};margin-top:6px;">Carryout / Use</div>'
            f'<div style="font-size:1.0rem;color:{DM_TEXT};">{cn_co_pct*100:.2f}%</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    with s3:
        sx_color = JSA_LT if "≥" in sx_section else COL_RED
        st.markdown(
            f'<div class="hl-card" style="text-align:left;">'
            f'<div class="hl-contract" style="margin-bottom:8px;">🫘 Nov Beans (SX)</div>'
            f'<div style="font-size:0.82rem;color:{DM_MUTED};">Jan 1 Price</div>'
            f'<div style="font-size:1.1rem;font-weight:600;color:{DM_TEXT};">${sx_jan1/100:.2f}/bu</div>'
            f'<div style="font-size:0.82rem;color:{DM_MUTED};margin-top:6px;">World C/O / Use</div>'
            f'<div style="font-size:1.0rem;color:{DM_TEXT};">{sx_co_pct*100:.2f}%'
            f' &nbsp;<span style="font-size:0.75rem;color:{sx_color};">'
            f'({sx_section})</span></div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    with s4:
        sn_color = {"No Crop Scare": JSA_LT, "SA Crop Problem": COL_GOLD, "Crop Scare": COL_RED}.get(sn_section, DM_MUTED)
        st.markdown(
            f'<div class="hl-card" style="text-align:left;">'
            f'<div class="hl-contract" style="margin-bottom:8px;">🫘 Jul Beans (SN)</div>'
            f'<div style="font-size:0.82rem;color:{DM_MUTED};">Jan 1 Price</div>'
            f'<div style="font-size:1.1rem;font-weight:600;color:{DM_TEXT};">${sn_jan1/100:.2f}/bu</div>'
            f'<div style="font-size:0.82rem;color:{DM_MUTED};margin-top:6px;">World C/O / Use</div>'
            f'<div style="font-size:1.0rem;color:{DM_TEXT};">{sn_co_pct*100:.2f}%'
            f' &nbsp;<span style="font-size:0.75rem;color:{sn_color};">'
            f'({sn_section})</span></div>'
            f'</div>',
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════
# ASSUMPTIONS TAB — all inputs (replaces sidebar)
# ══════════════════════════════════════════════════════════════
with tab_inp:
    st.markdown(
        f'<div class="sec-hdr">⚙️ Current Year Assumptions</div>'
        f'<div style="color:{DM_MUTED};font-size:0.82rem;margin-bottom:20px;">'
        f'Update these values to recalculate all indicated highs and lows across every contract.</div>',
        unsafe_allow_html=True,
    )

    col_a, col_b = st.columns(2)

    # ── CZ Dec Corn ──────────────────────────────────────────
    with col_a:
        st.markdown(
            f'<div class="input-card">'
            f'<div class="input-card-title">🌽 Dec Corn (CZ25)</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        cz_jan1_in = st.number_input(
            "Jan 1 Price (¢/bu)", value=float(cz_jan1), step=0.25,
            key="cz_jan1", format="%.2f",
            help="Futures price on January 1st of the marketing year, in cents/bushel.",
        )
        cz_prod_in = st.number_input(
            "US Production (Mil Bu)", value=float(cz_prod), step=50.0,
            key="cz_prod", format="%.0f",
        )
        cz_use_in = st.number_input(
            "US Usage (Mil Bu)", value=float(cz_use), step=50.0,
            key="cz_use", format="%.0f",
        )
        _cz_r = cz_prod_in / cz_use_in if cz_use_in else 0.0
        st.markdown(
            f'<span class="ratio-badge">Prod/Use: {_cz_r*100:.2f}%'
            f' &nbsp;·&nbsp; {"Prod ≥ Use" if _cz_r >= 1 else "Prod < Use"}</span>',
            unsafe_allow_html=True,
        )

    # ── CN Jul Corn ──────────────────────────────────────────
    with col_b:
        st.markdown(
            f'<div class="input-card">'
            f'<div class="input-card-title">🌽 Jul Corn (CN26)</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        cn_jan1_in = st.number_input(
            "Jan 1 Price (¢/bu)", value=float(cn_jan1), step=0.25,
            key="cn_jan1", format="%.2f",
        )
        cn_co_in = st.number_input(
            "Carryout % of Use", value=float(cn_co_raw), step=0.1,
            key="cn_co", format="%.1f",
            help="Enter as a percentage, e.g. 12.9 for 12.9%",
        )
        st.markdown(
            f'<span class="ratio-badge">C/O Ratio: {cn_co_in:.2f}%</span>',
            unsafe_allow_html=True,
        )

    st.markdown('<div style="height:18px;"></div>', unsafe_allow_html=True)
    col_c, col_d = st.columns(2)

    # ── SX Nov Soybeans ──────────────────────────────────────
    with col_c:
        st.markdown(
            f'<div class="input-card">'
            f'<div class="input-card-title">🫘 Nov Soybeans (SX25)</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        sx_jan1_in = st.number_input(
            "Jan 1 Price (¢/bu)", value=float(sx_jan1), step=0.25,
            key="sx_jan1", format="%.2f",
        )
        sx_co_in = st.number_input(
            "World C/O % of Use", value=float(sx_co_raw), step=0.1,
            key="sx_co", format="%.1f",
            help="Enter as a percentage, e.g. 28.4 for 28.4%",
        )
        st.selectbox(
            "Current Year Category",
            ["Prod ≥ Prev Yr Use", "Prod < Prev Yr Use"],
            key="sx_section",
            help="Is world production this year ≥ or < previous year's use? Highlights (★) matching historical years.",
        )
        st.markdown(
            f'<span class="ratio-badge">World C/O: {sx_co_in:.2f}%</span>',
            unsafe_allow_html=True,
        )

    # ── SN Jul Soybeans ──────────────────────────────────────
    with col_d:
        st.markdown(
            f'<div class="input-card">'
            f'<div class="input-card-title">🫘 Jul Soybeans (SN26)</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        sn_jan1_in = st.number_input(
            "Jan 1 Price (¢/bu)", value=float(sn_jan1), step=0.25,
            key="sn_jan1", format="%.2f",
        )
        sn_co_in = st.number_input(
            "World C/O % of Use", value=float(sn_co_raw), step=0.1,
            key="sn_co", format="%.1f",
            help="Enter as a percentage, e.g. 29.5 for 29.5%",
        )
        st.selectbox(
            "Current Year Category",
            ["No Crop Scare", "SA Crop Problem", "Crop Scare"],
            key="sn_section",
            help="Classify this marketing year — used to highlight (★) matching historical years in the ranked tables.",
        )
        st.markdown(
            f'<span class="ratio-badge">World C/O: {sn_co_in:.2f}%</span>',
            unsafe_allow_html=True,
        )

    st.markdown(
        f'<div style="margin-top:24px;color:{DM_MUTED};font-size:0.72rem;text-align:center;">'
        f'Data: USDA NASS &nbsp;·&nbsp; Model by JPSI &nbsp;·&nbsp; '
        f'Changes take effect immediately on all tabs.</div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════
# CZ TAB
# ══════════════════════════════════════════════════════════════
with tab_cz:
    st.markdown(
        '<div class="sec-hdr">Dec Corn (CZ) — Production as % of Same Year\'s Use</div>',
        unsafe_allow_html=True,
    )

    # Forecast tiles
    t1, t2 = st.columns(2)
    t1.markdown(headline_tile("CZ — Dec Corn", "Jan–Dec Indicated Low",  cz_hl_low,  cz_lp,  cz_t5_low,  "low"),  unsafe_allow_html=True)
    t2.markdown(headline_tile("CZ — Dec Corn", "Jan–Dec Indicated High", cz_hl_high, cz_hp, cz_t5_high, "high"), unsafe_allow_html=True)
    st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Jan 1 Price",  f"${cz_jan1/100:.2f}/bu")
    m2.metric("US Production", f"{cz_prod:,.0f} Mil Bu")
    m3.metric("US Usage",      f"{cz_use:,.0f} Mil Bu")
    m4.metric("Prod/Use Ratio", f"{cz_ratio*100:.2f}%",
              delta="Prod ≥ Use (Surplus)" if cz_ratio >= 1 else "Prod < Use (Deficit)",
              delta_color="normal" if cz_ratio >= 1 else "inverse")

    st.markdown("#### Historical Scatter: Prod/Use % vs Price as % of Jan 1")
    st.plotly_chart(
        make_scatter(D["cz_low"], D["cz_high"], "ratio",
                     "Prod/Use %", cz_ratio,
                     "Dec Corn: Production/Use vs Jan–Dec High/Low as % of Jan 1"),
        use_container_width=True,
    )

    st.markdown("#### Timing: Month of Annual High / Low (Historical Frequency)")
    st.plotly_chart(
        make_seasonality_chart(D["cz_low"], D["cz_high"],
                               "Dec Corn — Month When Annual High or Low Typically Occurs"),
        use_container_width=True,
    )

    col_l, col_h = st.columns(2)
    with col_l:
        tbl = build_display_table(D["cz_low"], "ratio", "Prod/Use %",
                                   cz_ratio, cz_jan1, "Jan–Dec Low (¢/bu)",
                                   current_section=cz_section)
        st.markdown(
            f'<div class="indicated-box">📉 <b>Lows</b> ranked by similarity &nbsp;|&nbsp; '
            f'Median indicated: <b style="color:{COL_LOW}">{fmt_price(cz_hl_low)}</b> '
            f'({fmt_pct(cz_lp)} of Jan 1) &nbsp;|&nbsp; '
            f'Top-5 comps: <b style="color:{COL_LOW}">{fmt_price(cz_t5_low)}</b>'
            f'&nbsp;·&nbsp; ★ = {cz_section}</div>',
            unsafe_allow_html=True,
        )
        st.dataframe(tbl, use_container_width=True, hide_index=True, height=480)

    with col_h:
        tbl = build_display_table(D["cz_high"], "ratio", "Prod/Use %",
                                   cz_ratio, cz_jan1, "Jan–Dec High (¢/bu)",
                                   current_section=cz_section)
        st.markdown(
            f'<div class="indicated-box">📈 <b>Highs</b> ranked by similarity &nbsp;|&nbsp; '
            f'Median indicated: <b style="color:{COL_HIGH}">{fmt_price(cz_hl_high)}</b> '
            f'({fmt_pct(cz_hp)} of Jan 1) &nbsp;|&nbsp; '
            f'Top-5 comps: <b style="color:{COL_HIGH}">{fmt_price(cz_t5_high)}</b>'
            f'&nbsp;·&nbsp; ★ = {cz_section}</div>',
            unsafe_allow_html=True,
        )
        st.dataframe(tbl, use_container_width=True, hide_index=True, height=480)


# ══════════════════════════════════════════════════════════════
# CN TAB
# ══════════════════════════════════════════════════════════════
with tab_cn:
    st.markdown(
        '<div class="sec-hdr">Jul Corn (CN) — Carryout as % of Use</div>',
        unsafe_allow_html=True,
    )

    # Forecast tiles
    t1, t2 = st.columns(2)
    t1.markdown(headline_tile("CN — Jul Corn", "Jan–Jul Indicated Low",  cn_hl_low,  cn_lp,  cn_t5_low,  "low"),  unsafe_allow_html=True)
    t2.markdown(headline_tile("CN — Jul Corn", "Jan–Jul Indicated High", cn_hl_high, cn_hp, cn_t5_high, "high"), unsafe_allow_html=True)
    st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)

    m1, m2 = st.columns(2)
    m1.metric("Jan 1 Price",      f"${cn_jan1/100:.2f}/bu")
    m2.metric("Carryout / Use",   f"{cn_co_pct*100:.2f}%")

    st.markdown("#### Historical Scatter: Carryout % vs Price as % of Jan 1")
    st.plotly_chart(
        make_scatter(D["cn_low"], D["cn_high"], "carryout_pct",
                     "Carryout % of Use", cn_co_pct,
                     "Jul Corn: Carryout/Use vs Jan–Jul High/Low as % of Jan 1"),
        use_container_width=True,
    )

    st.markdown("#### Timing: Month of Annual High / Low (Historical Frequency)")
    st.plotly_chart(
        make_seasonality_chart(D["cn_low"], D["cn_high"],
                               "Jul Corn — Month When Annual High or Low Typically Occurs"),
        use_container_width=True,
    )

    col_l, col_h = st.columns(2)
    with col_l:
        tbl = build_display_table(D["cn_low"], "carryout_pct", "Carryout %",
                                   cn_co_pct, cn_jan1, "Jan–Jul Low (¢/bu)")
        st.markdown(
            f'<div class="indicated-box">📉 <b>Lows</b> ranked by similarity &nbsp;|&nbsp; '
            f'Median indicated: <b style="color:{COL_LOW}">{fmt_price(cn_hl_low)}</b> '
            f'({fmt_pct(cn_lp)} of Jan 1) &nbsp;|&nbsp; '
            f'Top-5 comps: <b style="color:{COL_LOW}">{fmt_price(cn_t5_low)}</b></div>',
            unsafe_allow_html=True,
        )
        st.dataframe(tbl, use_container_width=True, hide_index=True, height=480)

    with col_h:
        tbl = build_display_table(D["cn_high"], "carryout_pct", "Carryout %",
                                   cn_co_pct, cn_jan1, "Jan–Jul High (¢/bu)")
        st.markdown(
            f'<div class="indicated-box">📈 <b>Highs</b> ranked by similarity &nbsp;|&nbsp; '
            f'Median indicated: <b style="color:{COL_HIGH}">{fmt_price(cn_hl_high)}</b> '
            f'({fmt_pct(cn_hp)} of Jan 1) &nbsp;|&nbsp; '
            f'Top-5 comps: <b style="color:{COL_HIGH}">{fmt_price(cn_t5_high)}</b></div>',
            unsafe_allow_html=True,
        )
        st.dataframe(tbl, use_container_width=True, hide_index=True, height=480)


# ══════════════════════════════════════════════════════════════
# SX TAB
# ══════════════════════════════════════════════════════════════
with tab_sx:
    st.markdown(
        '<div class="sec-hdr">Nov Soybeans (SX) — World Carryout as % of Same Year\'s Use</div>',
        unsafe_allow_html=True,
    )

    # Forecast tiles
    t1, t2 = st.columns(2)
    t1.markdown(headline_tile("SX — Nov Beans", "Jan–Nov Indicated Low",  sx_hl_low,  sx_lp,  sx_t5_low,  "low"),  unsafe_allow_html=True)
    t2.markdown(headline_tile("SX — Nov Beans", "Jan–Nov Indicated High", sx_hl_high, sx_hp, sx_t5_high, "high"), unsafe_allow_html=True)
    st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)

    m1, m2 = st.columns(2)
    m1.metric("Jan 1 Price",      f"${sx_jan1/100:.2f}/bu")
    m2.metric("World C/O / Use",  f"{sx_co_pct*100:.2f}%")

    st.markdown("#### Historical Scatter: World C/O % vs Price as % of Jan 1")
    st.plotly_chart(
        make_scatter(D["sx_low"], D["sx_high"], "carryout_pct",
                     "World C/O % of Use", sx_co_pct,
                     "Nov Soybeans: World C/O/Use vs Jan–Nov High/Low as % of Jan 1"),
        use_container_width=True,
    )

    st.markdown("#### Timing: Month of Annual High / Low (Historical Frequency)")
    st.plotly_chart(
        make_seasonality_chart(D["sx_low"], D["sx_high"],
                               "Nov Soybeans — Month When Annual High or Low Typically Occurs"),
        use_container_width=True,
    )

    col_l, col_h = st.columns(2)
    with col_l:
        tbl = build_display_table(D["sx_low"], "carryout_pct", "World C/O %",
                                   sx_co_pct, sx_jan1, "Jan–Nov Low (¢/bu)",
                                   current_section=sx_section)
        st.markdown(
            f'<div class="indicated-box">📉 <b>Lows</b> ranked by similarity &nbsp;|&nbsp; '
            f'Median indicated: <b style="color:{COL_LOW}">{fmt_price(sx_hl_low)}</b> '
            f'({fmt_pct(sx_lp)} of Jan 1) &nbsp;|&nbsp; '
            f'Top-5 comps: <b style="color:{COL_LOW}">{fmt_price(sx_t5_low)}</b>'
            f'&nbsp;·&nbsp; ★ = {sx_section}</div>',
            unsafe_allow_html=True,
        )
        st.dataframe(tbl, use_container_width=True, hide_index=True, height=480)

    with col_h:
        tbl = build_display_table(D["sx_high"], "carryout_pct", "World C/O %",
                                   sx_co_pct, sx_jan1, "Jan–Nov High (¢/bu)",
                                   current_section=sx_section)
        st.markdown(
            f'<div class="indicated-box">📈 <b>Highs</b> ranked by similarity &nbsp;|&nbsp; '
            f'Median indicated: <b style="color:{COL_HIGH}">{fmt_price(sx_hl_high)}</b> '
            f'({fmt_pct(sx_hp)} of Jan 1) &nbsp;|&nbsp; '
            f'Top-5 comps: <b style="color:{COL_HIGH}">{fmt_price(sx_t5_high)}</b>'
            f'&nbsp;·&nbsp; ★ = {sx_section}</div>',
            unsafe_allow_html=True,
        )
        st.dataframe(tbl, use_container_width=True, hide_index=True, height=480)


# ══════════════════════════════════════════════════════════════
# SN TAB
# ══════════════════════════════════════════════════════════════
with tab_sn:
    st.markdown(
        f'<div class="sec-hdr">Jul Soybeans (SN) — World Carryout as % of Use '
        f'<span style="color:{DM_MUTED};font-size:0.75rem;font-weight:400;">'
        f'· Current year: {SECTION_EMOJI.get(sn_section,"⚪")} {sn_section} '
        f'(★ in tables below)</span></div>',
        unsafe_allow_html=True,
    )

    # Forecast tiles
    t1, t2 = st.columns(2)
    t1.markdown(headline_tile("SN — Jul Beans", "Jan–Jul Indicated Low",  sn_hl_low,  sn_lp,  sn_t5_low,  "low"),  unsafe_allow_html=True)
    t2.markdown(headline_tile("SN — Jul Beans", "Jan–Jul Indicated High", sn_hl_high, sn_hp, sn_t5_high, "high"), unsafe_allow_html=True)
    st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)

    m1, m2, m3 = st.columns(3)
    m1.metric("Jan 1 Price",     f"${sn_jan1/100:.2f}/bu")
    m2.metric("World C/O / Use", f"{sn_co_pct*100:.2f}%")
    m3.metric("Current Category", sn_section)

    st.markdown("#### Historical Scatter: World C/O % vs Price as % of Jan 1 — All Three Categories")
    st.plotly_chart(
        make_scatter(D["sn_low"], D["sn_high"], "carryout_pct",
                     "World C/O % of Use", sn_co_pct,
                     "Jul Soybeans: World C/O/Use vs Jan–Jul High/Low as % of Jan 1"),
        use_container_width=True,
    )

    st.markdown("#### Timing: Month of Annual High / Low (Historical Frequency)")
    st.plotly_chart(
        make_seasonality_chart(D["sn_low"], D["sn_high"],
                               "Jul Soybeans — Month When Annual High or Low Typically Occurs"),
        use_container_width=True,
    )

    col_l, col_h = st.columns(2)
    with col_l:
        tbl = build_display_table(D["sn_low"], "carryout_pct", "World C/O %",
                                   sn_co_pct, sn_jan1, "Jan–Jul Low (¢/bu)",
                                   current_section=sn_section)
        st.markdown(
            f'<div class="indicated-box">📉 <b>Lows</b> ranked by similarity — all categories &nbsp;|&nbsp; '
            f'Median indicated: <b style="color:{COL_LOW}">{fmt_price(sn_hl_low)}</b> '
            f'({fmt_pct(sn_lp)} of Jan 1) &nbsp;|&nbsp; '
            f'Top-5 comps: <b style="color:{COL_LOW}">{fmt_price(sn_t5_low)}</b>'
            f'&nbsp;·&nbsp; ★ = {sn_section}</div>',
            unsafe_allow_html=True,
        )
        st.dataframe(tbl, use_container_width=True, hide_index=True, height=480)

    with col_h:
        tbl = build_display_table(D["sn_high"], "carryout_pct", "World C/O %",
                                   sn_co_pct, sn_jan1, "Jan–Jul High (¢/bu)",
                                   current_section=sn_section)
        st.markdown(
            f'<div class="indicated-box">📈 <b>Highs</b> ranked by similarity — all categories &nbsp;|&nbsp; '
            f'Median indicated: <b style="color:{COL_HIGH}">{fmt_price(sn_hl_high)}</b> '
            f'({fmt_pct(sn_hp)} of Jan 1) &nbsp;|&nbsp; '
            f'Top-5 comps: <b style="color:{COL_HIGH}">{fmt_price(sn_t5_high)}</b>'
            f'&nbsp;·&nbsp; ★ = {sn_section}</div>',
            unsafe_allow_html=True,
        )
        st.dataframe(tbl, use_container_width=True, hide_index=True, height=480)


# ══════════════════════════════════════════════════════════════
# SEASONALS TAB
# ══════════════════════════════════════════════════════════════
with tab_seas:
    hd_col, btn_col = st.columns([5, 1])
    with hd_col:
        st.markdown('<div class="sec-hdr">📈 Seasonal Price Patterns — % of Jan 1</div>',
                    unsafe_allow_html=True)
    with btn_col:
        st.markdown('<div style="height:20px;"></div>', unsafe_allow_html=True)
        if st.button("🔄 Refresh Data", key="refresh_price_btn", use_container_width=True,
                     help="Pulls the latest files from the SharePoint-synced folder and reloads the charts."):
            with st.spinner("Copying latest files…"):
                ok, msg = refresh_price_files()
            if ok:
                st.success(msg)
            else:
                st.error(msg)
            st.rerun()

    if PH is None:
        st.info(
            "📂 Price history files not found at the configured paths. "
            "This tab is available when running locally with access to the JSA network files. "
            "Update `CORN_PRICES_PATH` / `SOY_PRICES_PATH` at the top of `app.py` if the files have moved.",
            icon="ℹ️",
        )
    else:
        # Contract selector
        seas_contract = st.radio(
            "Contract",
            options=["🌽 Dec Corn (CZ25)", "🌽 Jul Corn (CN26)",
                     "🫘 Nov Soybeans (SX25)", "🫘 Jul Soybeans (SN26)"],
            horizontal=True,
            key="seas_contract",
        )

        st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)

        # ── Config per contract ───────────────────────────────
        SEAS_CONFIG = {
            "🌽 Dec Corn (CZ25)": dict(
                contracts=PH["corn"], prefix="ZCZ",
                current_year=2025,
                title="Dec Corn (CZ) — Annual Price as % of Jan 1  |  Each line = one delivery year",
                end_month=12,
            ),
            "🌽 Jul Corn (CN26)": dict(
                contracts=PH["corn"], prefix="ZCN",
                current_year=2026,
                title="Jul Corn (CN) — Annual Price as % of Jan 1  |  Each line = one delivery year",
                end_month=7,
            ),
            "🫘 Nov Soybeans (SX25)": dict(
                contracts=PH["soy"], prefix="ZSX",
                current_year=2025,
                title="Nov Soybeans (SX) — Annual Price as % of Jan 1  |  Each line = one delivery year",
                end_month=11,
            ),
            "🫘 Jul Soybeans (SN26)": dict(
                contracts=PH["soy"], prefix="ZSN",
                current_year=2026,
                title="Jul Soybeans (SN) — Annual Price as % of Jan 1  |  Each line = one delivery year",
                end_month=7,
            ),
        }

        cfg = SEAS_CONFIG[seas_contract]
        seas_df = build_seasonal_df(cfg["contracts"], cfg["prefix"])

        if seas_df.empty:
            st.warning("No price data found for this contract.")
        else:
            # Clip to marketing year window (Jan 1 → end month)
            seas_df = seas_df[seas_df["date"].dt.month <= cfg["end_month"]]

            # Summary metrics
            hist_df  = seas_df[seas_df["year"] != cfg["current_year"]]
            curr_df  = seas_df[seas_df["year"] == cfg["current_year"]]
            n_years  = hist_df["year"].nunique()
            avg_high = hist_df.groupby("year")["price_pct"].max().mean()
            avg_low  = hist_df.groupby("year")["price_pct"].min().mean()
            curr_last = curr_df["price_pct"].iloc[-1] if not curr_df.empty else None

            mc1, mc2, mc3, mc4 = st.columns(4)
            mc1.metric("Historical Years",    f"{n_years}")
            mc2.metric("Avg Seasonal High",   f"{avg_high:.1f}% of Jan 1")
            mc3.metric("Avg Seasonal Low",    f"{avg_low:.1f}% of Jan 1")
            if curr_last is not None:
                mc4.metric("Current Yr (Latest)", f"{curr_last:.1f}% of Jan 1")
            else:
                mc4.metric("Current Yr (Latest)", "No data yet")

            st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)

            # Seasonal overlay chart
            st.plotly_chart(
                make_seasonal_overlay(seas_df, cfg["current_year"], cfg["title"]),
                use_container_width=True,
            )

            # Year-by-year summary table
            with st.expander("📋 Year-by-Year Summary Table", expanded=False):
                summary_rows = []
                for yr, g in seas_df.groupby("year"):
                    jan1_raw = g[g["doy"] == g["doy"].min()]["close_raw"].iloc[0]
                    summary_rows.append({
                        "Year":          int(yr),
                        "Jan 1 ($/bu)":  f"${jan1_raw/100:.2f}",
                        "Seasonal High": f"{g['price_pct'].max():.1f}%",
                        "Seasonal Low":  f"{g['price_pct'].min():.1f}%",
                        "Range":         f"{g['price_pct'].max() - g['price_pct'].min():.1f}%",
                        "Current Yr":    "★" if yr == cfg["current_year"] else "",
                    })
                st.dataframe(
                    pd.DataFrame(summary_rows).sort_values("Year", ascending=False),
                    use_container_width=True, hide_index=True,
                )
